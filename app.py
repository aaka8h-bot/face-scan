import cv2
import face_recognition
import numpy as np
import os
import pickle
from cryptography.fernet import Fernet
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Folders setup
ENCODINGS_DIR = "encodings/"
ATTENDANCE_DIR = "attendance/"
os.makedirs(ENCODINGS_DIR, exist_ok=True)
os.makedirs(ATTENDANCE_DIR, exist_ok=True)

# Generate a secure encryption key (save this securely; in production, use a key management system)
def generate_key():
    key = Fernet.generate_key()
    with open("secret.key", "wb") as key_file:
        key_file.write(key)
    return key

# Load encryption key
def load_key():
    if not os.path.exists("secret.key"):
        return generate_key()
    return open("secret.key", "rb").read()

# Encrypt data
def encrypt_data(data, key):
    fernet = Fernet(key)
    return fernet.encrypt(data)

# Decrypt data
def decrypt_data(encrypted_data, key):
    fernet = Fernet(key)
    return fernet.decrypt(encrypted_data)

# Register a new user
def register_user(user_id):
    video_capture = cv2.VideoCapture(0)
    print(f"Registering user {user_id}. Look at the camera...")
    
    while True:
        ret, frame = video_capture.read()
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        if face_encodings:
            encoding = face_encodings[0]
            key = load_key()
            encrypted_encoding = encrypt_data(pickle.dumps(encoding), key)
            
            encoding_file = os.path.join(ENCODINGS_DIR, f"{user_id}.enc")
            with open(encoding_file, "wb") as f:
                f.write(encrypted_encoding)
            print(f"User {user_id} registered and encrypted data saved.")
            break
        
        cv2.imshow('Registration', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    video_capture.release()
    cv2.destroyAllWindows()

# Mark attendance in Excel
def mark_attendance(user_id):
    today = datetime.now().strftime("%Y-%m-%d")
    excel_file = os.path.join(ATTENDANCE_DIR, f"attendance_{today}.xlsx")
    
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["User ID", "Status", "Timestamp"])
        wb.save(excel_file)
    
    wb = load_workbook(excel_file)
    ws = wb.active
    timestamp = datetime.now().strftime("%H:%M:%S")
    ws.append([user_id, "Present", timestamp])
    wb.save(excel_file)
    print(f"Attendance marked for {user_id} at {timestamp}.")

# Attendance mode: Recognize and mark
def take_attendance():
    video_capture = cv2.VideoCapture(0)
    print("Attendance mode: Look at the camera...")
    
    key = load_key()
    known_encodings = {}
    for file in os.listdir(ENCODINGS_DIR):
        if file.endswith(".enc"):
            user_id = file.split(".")[0]
            with open(os.path.join(ENCODINGS_DIR, file), "rb") as f:
                encrypted = f.read()
            encoding = pickle.loads(decrypt_data(encrypted, key))
            known_encodings[user_id] = encoding
    
    while True:
        ret, frame = video_capture.read()
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
        
        for face_encoding in face_encodings:
            for user_id, known_encoding in known_encodings.items():
                match = face_recognition.compare_faces([known_encoding], face_encoding, tolerance=0.50)[0]
                if match:
                    mark_attendance(user_id)
                    video_capture.release()
                    cv2.destroyAllWindows()
                    return  # Exit after marking to avoid duplicates
        
        cv2.imshow('Attendance', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
    video_capture.release()
    cv2.destroyAllWindows()

# Main menu
if __name__ == "__main__":
    mode = input("Enter 'r' to register a user or 'a' for attendance: ").lower()
    if mode == 'r':
        user_id = input("Enter user ID (e.g., user1): ")
        register_user(user_id)
    elif mode == 'a':
        take_attendance()
    else:
        print("Invalid mode.")
